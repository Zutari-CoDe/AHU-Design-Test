"""
excel_export.py — Generate AHU_Design.xlsx with embedded psychrometric chart
=============================================================================
Produces the full calculation sheet matching the original Excel layout,
then embeds the live Plotly chart as a high-res PNG image at the bottom.
"""

import io
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter

# ── Color palette ─────────────────────────────────────────────────────────────
C_TITLE_BG   = "747474"
C_TITLE_FG   = "FFFFFF"
C_SECT_BG    = "86939F"
C_SECT_FG    = "FFFFFF"
C_SUB_BG     = "E6E9EB"
C_SUB_FG     = "0E2841"
C_PROC_BG    = "4EA72E"
C_PROC_FG    = "FFFFFF"
C_RED        = "FF0000"

# ── Style helpers ─────────────────────────────────────────────────────────────
def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _font(bold=False, color="000000", size=8, name="Arial"):
    return Font(bold=bold, color=color, size=size, name=name)

def _align(h="left", v="center", wrap=False, rot=0):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap, text_rotation=rot)

def _border_thin(top=False, bottom=False, left=False, right=False):
    s = Side(border_style="thin")
    return Border(
        top=s if top else None,
        bottom=s if bottom else None,
        left=s if left else None,
        right=s if right else None,
    )

def _apply_row_fill(ws, row, col_start, col_end, bg):
    for c in range(col_start, col_end + 1):
        ws.cell(row=row, column=c).fill = _fill(bg)

def _section_hdr(ws, row, num, label, bg=C_SECT_BG, fg=C_SECT_FG, ncols=17):
    _apply_row_fill(ws, row, 1, ncols, bg)
    c = ws.cell(row=row, column=1)
    c.value, c.font, c.fill = num, _font(True, fg), _fill(bg)
    c.alignment = _align(h="center")
    c = ws.cell(row=row, column=2)
    c.value, c.font, c.fill = label, _font(True, fg), _fill(bg)

def _sub_hdr(ws, row, label, ncols=17):
    _apply_row_fill(ws, row, 1, ncols, C_SUB_BG)
    c = ws.cell(row=row, column=2)
    c.value, c.font = label, _font(True, C_SUB_FG)

# ── Psychrometric formula templates ──────────────────────────────────────────
WEXLER = ('EXP(-5800.2206/(C{r}+273.15)+1.3914993'
          '-0.048640239*(C{r}+273.15)'
          '+0.000041764768*(C{r}+273.15)^2'
          '-0.000000014452093*(C{r}+273.15)^3'
          '+6.5459673*LN(C{r}+273.15))/1000')
MAGNUS = '0.61078*EXP((17.269*C{r})/(C{r}+237.3))'
PVSAT_WBT = ('IF(B{r}="","",EXP(-5800.2206/(D{r}+273.15)+1.3914993'
             '-0.048640239*(D{r}+273.15)'
             '+0.000041764768*(D{r}+273.15)^2'
             '-0.000000014452093*(D{r}+273.15)^3'
             '+6.5459673*LN(D{r}+273.15))/1000)')

MOIST_FMTS = {
    'C': '0.0"°CDB"', 'D': '0.0"°CWB"', 'E': '0.00%', 'F': '0.00"g/kg"',
    'G': '0.0"kJ/kg"', 'H': '0.0"°C"',  'I': '0.000"kPa"','J': '0.000"kPa"',
    'K': '0.000"kPa"', 'L': '0.000"kg/m3"', 'M': '0.000"kJ/kg.K"', 'N': '0.0',
}

def _moist_row(ws, r, name, tdb, twb, ashrae=True, red_inputs=True):
    ws.cell(row=r, column=2).value = name
    ws.cell(row=r, column=2).font = _font()

    for col, val, red in [(3, tdb, red_inputs), (4, twb, red_inputs)]:
        c = ws.cell(row=r, column=col)
        c.value = val
        c.font = _font(color=C_RED if red else "000000")
        c.alignment = _align(h="center")

    pvsat_dbt = WEXLER.format(r=r) if ashrae else MAGNUS.format(r=r)
    pv_f = (f'IF(B{r}="","",J{r}-($D$17*0.000665*(C{r}-D{r})))' if ashrae
            else f'IF(B{r}="","",J{r}-0.000665*$D$17*(C{r}-D{r}))')

    formulas = {
        'E': f'=IF(B{r}="","",K{r}/I{r})',
        'F': f'=IFERROR(622*K{r}/($D$17-K{r}),"")' ,
        'G': f'=IFERROR(IF(B{r}="","",1.006*C{r}+N{r}/1000*(2501+1.86*C{r})),"")',
        'H': (f'=IFERROR(237.3*(LN(E{r})+(17.27*C{r})/(237.3+C{r}))'
              f'/(17.27-(LN(E{r})+(17.27*C{r})/(237.3+C{r}))),"")'),
        'I': f'=IF(B{r}="","",{pvsat_dbt})',
        'J': f'={PVSAT_WBT.format(r=r)}',
        'K': f'={pv_f}',
        'L': f'=IFERROR(IF(B{r}="","",$D$17/(287.05*(273+C{r})))*1000,"")',
        'M': f'=IF(B{r}="","",((1005+0.1*C{r}+0.000025*C{r}^2)*($D$17/101.325)^-0.07)/1000)',
        'N': f'=IFERROR(IF(B{r}="","",(0.62198*K{r})/($D$17-K{r})*1000),"")',
    }
    for col_ltr, formula in formulas.items():
        c = ws[f'{col_ltr}{r}']
        c.value = formula
        c.font = _font()
        c.alignment = _align(h="center")

    for col_ltr, fmt in MOIST_FMTS.items():
        ws[f'{col_ltr}{r}'].number_format = fmt


def _proc_pair(ws, r_in, r_out, name, in_state, out_state, flow_ref='$D$15', note=''):
    ws.merge_cells(f'B{r_in}:B{r_out}')
    c = ws.cell(row=r_in, column=2)
    c.value, c.font = name, _font(color=C_RED)
    c.alignment = _align(h="center", wrap=True)

    for r, state, io in [(r_in, in_state, 'In'), (r_out, out_state, 'Out')]:
        ws.cell(row=r, column=3).value = io
        ws.cell(row=r, column=3).font = _font()
        ws.cell(row=r, column=3).alignment = _align(h="center")
        ws.cell(row=r, column=4).value = state
        ws.cell(row=r, column=4).font = _font(color=C_RED)
        ws.cell(row=r, column=4).alignment = _align(h="center")

        tdb_f = f'=IFERROR(INDEX($C$27:$N$51,MATCH($D{r},$B$27:$B$51,0),1),"")'
        w_f   = f'=IFERROR(INDEX($C$27:$N$51,MATCH($D{r},$B$27:$B$51,0),4),"")'
        mdot_f= f'={flow_ref}*(INDEX($C$27:$N$51,MATCH($D{r},$B$27:$B$51,0),10))'

        for col, val, fmt in [
            (5, tdb_f,  '0.0"°C"'),
            (6, w_f,    '0.00"g/kg"'),
            (7, mdot_f, '0.000"kg/s"'),
        ]:
            c = ws.cell(row=r, column=col)
            c.value, c.number_format = val, fmt
            c.alignment = _align(h="center")

    # Heat calcs on in-row — col 4=W(g/kg), 5=h in $C$27:$N$51
    ws[f'H{r_in}'].value = (
        f'=G{r_out}*(INDEX($C$27:$N$51,MATCH(D{r_in},$B$27:$B$51,0),4)'
        f'-INDEX($C$27:$N$51,MATCH(D{r_out},$B$27:$B$51,0),4))'
    )
    ws[f'H{r_in}'].number_format = '0.0"g/s"'
    ws[f'H{r_in}'].alignment = _align(h="center")

    ws[f'I{r_in}'].value = (
        f'=G{r_out}*INDEX($C$27:$N$51,MATCH($D{r_in},$B$27:$B$51,0),11)'
        f'*(INDEX($C$27:$N$51,MATCH($D{r_out},$B$27:$B$51,0),1)'
        f'-INDEX($C$27:$N$51,MATCH($D{r_in},$B$27:$B$51,0),1))'
    )
    ws[f'I{r_in}'].number_format = '0.00"kW"'
    ws[f'I{r_in}'].alignment = _align(h="center")

    ws[f'J{r_in}'].value = f'=K{r_in}-I{r_in}'
    ws[f'J{r_in}'].number_format = '0.00"kW"'
    ws[f'J{r_in}'].alignment = _align(h="center")

    ws[f'K{r_in}'].value = (
        f'=-G{r_out}*(INDEX($C$27:$N$51,MATCH(D{r_in},$B$27:$B$51,0),5)'
        f'-INDEX($C$27:$N$51,MATCH(D{r_out},$B$27:$B$51,0),5))'
    )
    ws[f'K{r_in}'].font = _font(bold=True)
    ws[f'K{r_in}'].number_format = '0.00"kW"'
    ws[f'K{r_in}'].alignment = _align(h="center")

    ws[f'L{r_in}'].value = f'=IFERROR(I{r_in}/K{r_in},"")'
    ws[f'L{r_in}'].number_format = '0.000'
    ws[f'L{r_in}'].alignment = _align(h="center")

    if note:
        ws.cell(row=r_in, column=13).value = note
        ws.cell(row=r_in, column=13).font = _font(size=8)


# ── Main builder ──────────────────────────────────────────────────────────────

def build_excel(inp: dict, fig: go.Figure, states: dict = None, P: float = None) -> bytes:
    """
    Build the full AHU Design Excel workbook and return as bytes.

    Parameters
    ----------
    inp    : dict         — full input dict (same as Flask routes receive)
    fig    : go.Figure    — the Plotly psychrometric chart (kept for API compat)
    states : dict         — AirState objects keyed by name (for chart rendering)
    P      : float        — air pressure in Pa
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "AHU DESIGN"

    # ── Column widths ──────────────────────────────────────────────────────
    col_widths = {
        'A': 2.66, 'B': 19.89, 'C': 8.66, 'D': 14.89, 'E': 8.11,
        'F': 8.11,  'G': 8.66,  'H': 8.66, 'I': 13.0,  'J': 8.11,
        'K': 9.44,  'L': 9.0,   'M': 10.11,'N': 5.89,  'O': 10.55,
        'P': 24.0,  'Q': 16.66,
    }
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

    # ── Row heights ────────────────────────────────────────────────────────
    for row, h in {1:12,2:12,5:16,10:16,20:16,21:87,22:16,27:15,
                   29:15,30:15,31:15,39:15,40:20,52:16,53:51,
                   57:20,61:20,67:31,68:31,69:31}.items():
        ws.row_dimensions[row].height = h

    # ── Rows 1-2: Notes ───────────────────────────────────────────────────
    ws['B1'].value = 'NOTE: ONLY MODIFY RED VALUES, THE OTHERS ARE CALCULATED'
    ws['B1'].font = _font(True, C_RED)
    ws['B2'].value = 'ADJUST VIEW — select chart, choose format axis, update min/max as required'
    ws['B2'].font = _font(True, C_RED)

    # ── Row 5: Title bar ──────────────────────────────────────────────────
    ws.merge_cells('A5:P5')
    ws['A5'].value = 'CALCULATION SHEET'
    ws['A5'].font = _font(True, C_TITLE_FG)
    ws['A5'].fill = _fill(C_TITLE_BG)
    ws['A5'].alignment = _align(h="center")
    for c in range(1, 17):
        ws.cell(row=5, column=c).fill = _fill(C_TITLE_BG)

    # ── Section 1: Global Inputs ──────────────────────────────────────────
    _section_hdr(ws, 10, 1, 'GLOBAL INPUTS')

    for r, label in [(6,'Project'),(7,'Project number'),(8,'Revision'),(9,'Author')]:
        ws.cell(row=r, column=2).value = label
        ws.cell(row=r, column=2).font = _font(True)

    project_vals = {
        6: inp.get('project', 'PROJECT NAME'),
        7: inp.get('project_number', 'PROJ-001'),
        8: inp.get('revision', 'REV 0'),
        9: inp.get('author', 'ENGINEER'),
    }
    for r, val in project_vals.items():
        ws.cell(row=r, column=4).value = val
        ws.cell(row=r, column=4).font = _font()

    ws['F9'].value = '*input parameters are indicated in red'
    ws['F9'].font = _font(color=C_RED)

    global_rows = [
        (11, 'FACILITY LOCATION',            inp.get('city',''),              True,  'General'),
        (12, 'SPECIFIC DATA HALL IT LOAD',   inp.get('it_load', 1500),        True,  '0'),
        (13, 'DATA HALL SENSIBLE LOAD',       '=D12+D12*0.055',               False, '0.0'),
        (14, 'DATA HALL AIR FLOW',            '=D13/(C29-C28)/M29',           False, '0.000'),
        (15, 'AHU Off coil air flow calc.',   inp.get('ahu_vol_flow', 1.605),  True,  '0.000'),
        (16, 'Height above sea level',        inp.get('altitude', 22),         True,  '0'),
        (17, 'Air pressure at altitude',      '=101.325*(1-0.0000225577*$D$16)^5.25588', False, '0.000'),
        (18, 'Fan electrical load',           '=D15*600/1000',                False, '0.000'),
        (19, 'Fan delta T',                   '=$D$18/($D$15*$M$37)',          False, '0.000'),
    ]
    for r, label, val, red, fmt in global_rows:
        ws.cell(row=r, column=2).value = label
        ws.cell(row=r, column=2).font = _font()
        c = ws.cell(row=r, column=4)
        c.value = val
        c.font = _font(color=C_RED if red else "000000")
        c.alignment = _align(h="center")
        c.number_format = fmt

    ws['E18'].value = 'est'
    ws['F11'].value = 'WEATHER STATION - ADD WEATHER AT END'
    ws['F13'].value = 'For calculating crah temps: 5.5% aux load factor added to IT load'

    # ── Section 2: Moist Air States ───────────────────────────────────────
    _section_hdr(ws, 20, 2, 'MOIST AIR STATES')

    # Row 21: column headers (rotated)
    hdrs_21 = [
        (2,'Point name'),(3,'Dry bulb temp (Tdb)'),(4,'Wet bulb temp (Twb)'),
        (5,'Relative Humidity (RH)'),(6,'Absolute humidity (W)'),(7,'Enthalpy (h)'),
        (8,'Dew point (Tdp)'),(9,'Saturation vapour pressure @DBT'),
        (10,'Saturation vapour pressure @WBT'),(11,'Actual vapour pressure (Pv)'),
        (12,'Density (ρ)'),(13,'Specific heat capacity (Cp)'),(14,'Humidity ratio (W)'),
        (15,'Notes'),
    ]
    for col, hdr in hdrs_21:
        cell = ws.cell(row=21, column=col)
        cell.value = hdr
        cell.font = _font(True)
        cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                   wrap_text=True, text_rotation=90)
        cell.border = Border(
            top=Side(border_style="thin"),
            left=Side(border_style="thin") if col > 2 else None,
        )

    ws.cell(row=21, column=15).value = 'Notes / ASHRAE Limits'
    ws.cell(row=21, column=15).font = _font(True)
    ws.cell(row=21, column=15).alignment = Alignment(
        horizontal="center", vertical="bottom", wrap_text=True, text_rotation=90)

    # ── 2.1 ASHRAE A1 zone ─────────────────────────────────────────────
    _sub_hdr(ws, 22, 'ASHRAE COLD AISLE ZONE FOR DATA CENTRES')
    ws.cell(row=22, column=16).value = 'Ashrae dp'
    ws.cell(row=22, column=16).font = _font(True, C_SUB_FG)
    ws.cell(row=22, column=17).value = 'Ashrae RH'
    ws.cell(row=22, column=17).font = _font(True, C_SUB_FG)

    ashrae_data = [
        (23, 'Ash 18 constant temp', inp.get('ash_18_low_tdb', 18),  inp.get('ash_twb_low', 6.4),    '>-9',  '-'),
        (24, 'Ash 18 constant temp', inp.get('ash_18_high_tdb', 18), inp.get('ash_twb_high', 14.4), '<15',  '<60% (CAN BE 70% IF CORROSION COUPON OK)'),
        (25, 'Ash 27 constant temp', inp.get('ash_27_low_tdb', 27),  inp.get('ash_twb_27_low', 13.2),  '<15',  '<60% (CAN BE 70% IF CORROSION COUPON OK)'),
        (26, 'Ash 27 constant temp', inp.get('ash_27_high_tdb', 27), inp.get('ash_twb_27_high', 10.27),'<-9', '-'),
    ]
    for r, name, tdb, twb, dp, rh_note in ashrae_data:
        _moist_row(ws, r, name, tdb, twb, ashrae=True)
        ws[f'O{r}'].value = dp
        ws[f'P{r}'].value = rh_note

    # ── 2.2 Control conditions ─────────────────────────────────────────
    _section_hdr(ws, 27, 2, 'CONTROL CONDITIONS')
    _moist_row(ws, 28, 'CRAH OFF COIL', inp.get('crah_off_tdb',25), inp.get('crah_off_twb',16.5))
    ws['B28'].alignment = _align(h="center")
    ws['O28'].value = 'CRAH unit off coil — ensures no load on CRAH at start'
    _moist_row(ws, 29, 'CRAH ON COIL',  inp.get('crah_on_tdb',36),  inp.get('crah_on_twb',19.8))
    ws['B29'].alignment = _align(h="center")

    # ── 2.3 Outdoor inlet ──────────────────────────────────────────────
    _section_hdr(ws, 30, 3, 'OUTDOOR INLET CONDITIONS')
    outdoor = [
        (31, 'MAX OAT (N=20)',  inp.get('oat_n20_tdb',49.2),    inp.get('oat_n20_twb',32.9),    'Adjust WB to meet humidity as needed'),
        (32, 'MAX OAE (0.4%)',  inp.get('oat_04e_tdb',35.2),    inp.get('oat_04e_twb',30.75),   ''),
        (33, 'MAX OAH (0.4%)',  inp.get('oat_04h_tdb',33.6),    inp.get('oat_04h_twb',30.2),    ''),
        (34, 'MIN OAT (N=20)',  inp.get('oat_min_n20_tdb',7.3), inp.get('oat_min_n20_twb',3.6), ''),
        (35, 'MIN OAH (99.6%)', inp.get('oat_min_04h_tdb',31.1),inp.get('oat_min_04h_twb',14.7),''),
    ]
    for r, name, tdb, twb, note in outdoor:
        _moist_row(ws, r, name, tdb, twb, ashrae=False)
        ws[f'B{r}'].alignment = _align(h="center")
        if note:
            ws[f'O{r}'].value = note

    # ── 2.4 AHU off-coil ───────────────────────────────────────────────
    _sub_hdr(ws, 36, 'AHU OFF COIL CONDITIONS')
    oc_data = [
        (37, 'OFF COIL MAX COOL',             inp.get('oc_cool_tdb',12.9), '=C37',  'MATCH CRAH UNIT DEW POINT'),
        (38, 'OFF COIL ENTHALPY',             inp.get('oc_enth_tdb',14.55),'=C38',  'Target 42-46 kJ/kg enthalpy'),
        (39, 'OFF COIL DEHUM',                inp.get('oc_dehum_tdb',15),  '=C39',  'Match CRAH DP or increase as long as CRAH can dehumidify'),
        (40, 'OFF COIL HEATING FOR HUMIDIFICATION', inp.get('oc_heat_tdb',36), inp.get('oc_heat_twb',15.82), 'MATCH min oah dew point'),
        (41, 'OFF COIL HUMIDIFICATION',       15,  '=C41', ''),
    ]
    for r, name, tdb, twb, note in oc_data:
        _moist_row(ws, r, name, tdb, twb, ashrae=False)
        ws[f'B{r}'].alignment = _align(h="center")
        ws[f'D{r}'].value = twb
        ws[f'D{r}'].font = _font()
        if note:
            ws[f'O{r}'].value = note

    # ── 2.5 Return air ─────────────────────────────────────────────────
    _section_hdr(ws, 42, 4, 'RETURN AIR CONDITIONS')
    _moist_row(ws, 43, 'RA CONDITION', inp.get('ra_tdb',35), inp.get('ra_twb',25), ashrae=False)
    ws['B43'].alignment = _align(h="center")

    # ── 2.6 Energy recovery ────────────────────────────────────────────
    _sub_hdr(ws, 44, 'ENERGY RECOVERY')
    _moist_row(ws, 45, 'E-WHEEL',       '=C43', '=D43', ashrae=False)
    ws['B45'].alignment = _align(h="center")
    _moist_row(ws, 46, 'RUN AROUND COIL', 0, 0, ashrae=False)
    ws['B46'].alignment = _align(h="center")

    for r in [48, 49, 50, 51]:
        _moist_row(ws, r, '', 0, 0, ashrae=False)

    ws.cell(row=48, column=2).value = 'Heating coil LAT'
    ws.cell(row=48, column=3).value = 18.5
    ws.cell(row=48, column=4).value = 18.54

    # ── Section 3: Psychrometric Evaluation ──────────────────────────────
    _section_hdr(ws, 52, 2, 'PSYCHROMETRIC EVALUATION')

    proc_hdrs = [
        (2,'Process name'),(3,'State IN'),(4,'State OUT'),
        (5,'Tdb (°C)'),(6,'W (g/kg)'),(7,'ṁ (kg/s)'),(8,'Humidif. (g/s)'),
        (9,'Q_sens (kW)'),(10,'Q_lat (kW)'),(11,'Q_total (kW)'),(12,'SHR'),
    ]
    for col, hdr in proc_hdrs:
        cell = ws.cell(row=53, column=col)
        cell.value = hdr
        cell.font = Font(bold=True, color=C_PROC_FG, size=8, name='Arial')
        cell.fill = _fill(C_PROC_BG)
        cell.alignment = Alignment(horizontal="center", vertical="bottom",
                                   wrap_text=True, text_rotation=90)
    for col in range(1, 13):
        ws.cell(row=53, column=col).fill = _fill(C_PROC_BG)

    _proc_pair(ws, 54, 55, 'SENSIBLE COOLING TO CRAH OFF COIL',
               'MAX OAT (N=20)', 'MAX OAT (N=20)')
    _proc_pair(ws, 56, 57, 'TOTAL OAT COOLING',
               'MAX OAT (N=20)', 'OFF COIL MAX COOL')
    _proc_pair(ws, 58, 59, 'MAX DEHUMIDIFICATION',
               'MAX OAH (0.4%)', 'OFF COIL DEHUM')
    _proc_pair(ws, 60, 61, 'MAX ENTHALPY COOLING',
               'MAX OAE (0.4%)', 'OFF COIL ENTHALPY')
    _proc_pair(ws, 62, 63, 'CRAH LOAD',
               'CRAH ON COIL', 'CRAH OFF COIL', flow_ref='$D$14',
               note='match sensible load above')
    _proc_pair(ws, 64, 65, 'DOAS FAN HEAT LOAD',
               'OFF COIL MAX COOL', 'CRAH OFF COIL')
    _proc_pair(ws, 66, 67, 'WINTER HEATING FOR HUMIDIFICATION',
               'MIN OAH (99.6%)', 'OFF COIL HEATING FOR HUMIDIFICATION')
    _proc_pair(ws, 68, 69, 'HUMIDIFICATION',
               'OFF COIL HEATING FOR HUMIDIFICATION',
               'OFF COIL HEATING FOR HUMIDIFICATION')

    # ── Psychrometric Chart Image ─────────────────────────────────────────
    chart_start_row = 72

    # Spacer / header row
    _section_hdr(ws, chart_start_row, '', 'PSYCHROMETRIC CHART', bg=C_TITLE_BG, fg=C_TITLE_FG)
    ws.row_dimensions[chart_start_row].height = 16

    # Render chart as PNG using matplotlib (no Chrome dependency)
    if states is not None and P is not None:
        from chart_png import render_chart_png
        img_bytes = render_chart_png(inp, states, P)
    else:
        # Fallback: blank placeholder
        img_bytes = _blank_chart_png()

    img_stream = io.BytesIO(img_bytes)
    xl_img = XLImage(img_stream)
    xl_img.width  = 1200
    xl_img.height = 600
    ws.add_image(xl_img, f'B{chart_start_row + 1}')

    for r in range(chart_start_row + 1, chart_start_row + 35):
        ws.row_dimensions[r].height = 20

    # ── Save to temp file, recalculate formulas, read back ────────────────
    import tempfile, os, subprocess, json, sys

    with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        tmp_path = tmp.name

    try:
        wb.save(tmp_path)

        # Run LibreOffice recalc to populate all formula values
        scripts_dir = os.path.join(os.path.dirname(__file__), "scripts")
        recalc_script = os.path.join(scripts_dir, "recalc.py")
        result = subprocess.run(
            [sys.executable, recalc_script, tmp_path, "60"],
            capture_output=True, text=True, timeout=90
        )
        # Log any errors but don't fail — file is still usable with formulas
        try:
            info = json.loads(result.stdout)
            if info.get("total_errors", 0) > 0:
                print(f"[excel_export] recalc errors: {info['error_summary']}")
        except Exception:
            pass

        with open(tmp_path, "rb") as f:
            return f.read()
    finally:
        try:
            os.unlink(tmp_path)
        except Exception:
            pass


def _blank_chart_png() -> bytes:
    """Fallback: a plain white 1200x600 PNG."""
    import matplotlib
    matplotlib.use('Agg')
    import matplotlib.pyplot as plt
    fig, ax = plt.subplots(figsize=(16, 8))
    ax.text(0.5, 0.5, "Psychrometric Chart\n(states not available)",
            ha='center', va='center', transform=ax.transAxes, fontsize=14, color='#888')
    buf = io.BytesIO()
    fig.savefig(buf, format='png', dpi=100, bbox_inches='tight', facecolor='white')
    plt.close(fig)
    buf.seek(0)
    return buf.read()
    """Convert the dark-theme chart to a clean white-background version for Excel."""
    import copy, plotly.graph_objects as go

    wfig = go.Figure(fig)  # shallow copy — we'll update layout

    # Color mappings: dark → light equivalents
    dark_to_light = {
        "#0f1117": "#ffffff",
        "#1e2230": "#f0f0f0",
        "rgba(150,150,150,0.35)": "rgba(180,180,180,0.5)",
        "rgba(100,180,255,0.20)": "rgba(30,100,200,0.15)",
        "rgba(200,200,100,0.20)": "rgba(150,150,0,0.15)",
        "rgba(46,204,113,0.10)":  "rgba(46,204,113,0.12)",
        "#ffffff": "#1a1a2e",     # saturation curve: white → dark navy
        "#cccccc": "#333333",
        "#2a2d3e": "#cccccc",
        "rgba(15,17,23,0.85)":    "rgba(255,255,255,0.9)",
        "#00c3ff": "#0066aa",
        "#666": "#555555",
        "#336699": "#1a5276",
    }

    def remap(color):
        if color is None:
            return color
        return dark_to_light.get(str(color), color)

    # Update traces
    new_traces = []
    for trace in wfig.data:
        td = trace.to_plotly_json()

        # Line color
        if 'line' in td and td['line']:
            td['line']['color'] = remap(td['line'].get('color'))

        # Marker
        if 'marker' in td and td['marker']:
            td['marker']['color'] = remap(td['marker'].get('color'))
            if 'line' in td['marker']:
                mc = td['marker']['line'].get('color')
                td['marker']['line']['color'] = '#333333' if mc == '#fff' else remap(mc)

        # Fill color
        if 'fillcolor' in td:
            td['fillcolor'] = remap(td.get('fillcolor'))

        # Text font color
        if 'textfont' in td and td['textfont']:
            td['textfont']['color'] = remap(td['textfont'].get('color'))

        # Rebuild trace of same type
        trace_type = type(trace)
        new_traces.append(trace_type(**{k: v for k, v in td.items() if k != 'type'}))

    wfig.data = []
    for t in new_traces:
        wfig.add_trace(t)

    # Layout overrides
    wfig.update_layout(
        template="plotly_white",
        paper_bgcolor="#ffffff",
        plot_bgcolor="#ffffff",
        font=dict(family="Arial", size=11, color="#1a1a2e"),
        title=dict(
            font=dict(color="#0066aa", size=13),
        ),
        xaxis=dict(
            gridcolor="#e8e8e8",
            linecolor="#cccccc",
            tickcolor="#666666",
            title_font=dict(color="#333333"),
        ),
        yaxis=dict(
            gridcolor="#e8e8e8",
            linecolor="#cccccc",
            tickcolor="#666666",
            title_font=dict(color="#333333"),
        ),
        legend=dict(
            bgcolor="rgba(255,255,255,0.95)",
            bordercolor="#cccccc",
            borderwidth=1,
            font=dict(color="#333333", size=10),
        ),
    )

    return wfig
